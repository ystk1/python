#!/usr/local/bin/python
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import pyexcel
import getopt
import sys


def main():
    xls_path = sys.argv[1]
    xlsx_path = xls_path + 'x'

    pyexcel.save_book_as(file_name=xls_path,
                         dest_file_name=xlsx_path)
    
    wbook = load_workbook(xlsx_path, data_only = True)
    
    for sheet_name in wbook.get_sheet_names():
        print '## SHEET_NAME=' + sheet_name
        wsheet = wbook.get_sheet_by_name(sheet_name)

        cell = wsheet.cell(row=8,column=11)
        val = cell.value
        val_2 = cell.internal_value
        print "%s,%s" % (val,val_2)

        return

if __name__ == '__main__':
    main()