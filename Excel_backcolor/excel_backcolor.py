import openpyxl as px
from openpyxl.styles import Border, Side
 
book = px.Workbook()
book.save('sample.xlsx')
 
wb = px.load_workbook('sample.xlsx')
ws = wb['Sheet']
 
## 背景色を変更
fill = px.styles.PatternFill(patternType='solid',fgColor='D1FE7B',bgColor='D1FE7B')
 
## セルの指定
ws['A5'].fill = fill
 
wb.save('sample.xlsx')
print('セルの背景色変更完了')